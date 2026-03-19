"""
Report service for advanced analytics, custom report builder, and exports.
"""
import io
import json
import logging
from datetime import datetime, timezone

from sqlalchemy import func

from models import db
from models_crm import Deal, DealStage, Report, ReportSchedule
from services.analytics_service import AnalyticsService

logger = logging.getLogger(__name__)


class ReportService:
    """Service layer for saved reports, custom reports, and report exports."""

    SYSTEM_REPORT_TYPES = {'pipeline', 'forecast', 'win_loss', 'cycle', 'stage_conversion'}
    CUSTOM_DIMENSIONS = {'stage', 'owner', 'status', 'month'}
    CUSTOM_METRICS = {'count', 'sum', 'average'}

    @staticmethod
    def create_report(workspace_id, created_by, name, report_type, config=None):
        """Create a saved report definition."""
        if not name or not str(name).strip():
            raise ValueError('Report name is required')

        if report_type not in ReportService.SYSTEM_REPORT_TYPES and report_type != 'custom':
            raise ValueError('Invalid report type')

        report = Report(
            workspace_id=workspace_id,
            name=str(name).strip(),
            report_type=report_type,
            config_json=json.dumps(config or {}),
            created_by=created_by,
        )

        try:
            db.session.add(report)
            db.session.commit()
            return report
        except Exception as exc:
            db.session.rollback()
            logger.error('Failed to create report: %s', exc)
            raise

    @staticmethod
    def list_reports(workspace_id):
        """List saved reports for workspace."""
        reports = Report.query.filter_by(workspace_id=workspace_id).order_by(Report.created_at.desc()).all()
        return [ReportService.serialize_report(row) for row in reports]

    @staticmethod
    def get_report(report_id, workspace_id):
        """Get report by id with workspace isolation."""
        return Report.query.filter_by(id=report_id, workspace_id=workspace_id).first()

    @staticmethod
    def run_report(report, workspace_id):
        """Execute a report and return normalized output."""
        config = ReportService._safe_config(report.config_json)

        if report.report_type in {'pipeline', 'forecast'}:
            return {
                'report_type': report.report_type,
                'generated_at': datetime.utcnow().isoformat(),
                'data': AnalyticsService.get_pipeline_distribution(workspace_id),
            }

        if report.report_type == 'win_loss':
            return {
                'report_type': report.report_type,
                'generated_at': datetime.utcnow().isoformat(),
                'data': AnalyticsService.get_win_loss_ratio(workspace_id),
            }

        if report.report_type == 'cycle':
            days = int(config.get('days', 90) or 90)
            return {
                'report_type': report.report_type,
                'generated_at': datetime.utcnow().isoformat(),
                'data': AnalyticsService.get_sales_cycle_duration(workspace_id, days),
            }

        if report.report_type == 'stage_conversion':
            return {
                'report_type': report.report_type,
                'generated_at': datetime.utcnow().isoformat(),
                'data': AnalyticsService.get_stage_conversion_rate(workspace_id),
            }

        if report.report_type == 'custom':
            return {
                'report_type': report.report_type,
                'generated_at': datetime.utcnow().isoformat(),
                'data': ReportService.run_custom_report(workspace_id, config),
            }

        raise ValueError('Unsupported report type')

    @staticmethod
    def run_custom_report(workspace_id, config):
        """Run custom builder report query with limited dimensions/metrics."""
        dimension = (config or {}).get('dimension', 'stage')
        metric = (config or {}).get('metric', 'count')

        if dimension not in ReportService.CUSTOM_DIMENSIONS:
            raise ValueError('Invalid dimension')
        if metric not in ReportService.CUSTOM_METRICS:
            raise ValueError('Invalid metric')

        base_query = db.session.query()

        if dimension == 'stage':
            dim_expr = DealStage.name.label('dimension')
            base_query = base_query.join(DealStage, Deal.stage_id == DealStage.id)
        elif dimension == 'owner':
            from models import User
            dim_expr = User.name.label('dimension')
            base_query = base_query.join(User, Deal.owner_id == User.id)
        elif dimension == 'status':
            dim_expr = Deal.status.label('dimension')
        else:
            dim_expr = func.strftime('%Y-%m', Deal.created_at).label('dimension')

        if metric == 'count':
            metric_expr = func.count(Deal.id).label('metric')
        elif metric == 'sum':
            metric_expr = func.coalesce(func.sum(Deal.value), 0).label('metric')
        else:
            metric_expr = func.coalesce(func.avg(Deal.value), 0).label('metric')

        query = base_query.with_entities(dim_expr, metric_expr).filter(
            Deal.workspace_id == workspace_id
        ).group_by(dim_expr).order_by(metric_expr.desc())

        rows = query.all()
        return {
            'dimension': dimension,
            'metric': metric,
            'rows': [
                {'dimension': row.dimension or 'N/A', 'value': float(row.metric or 0)}
                for row in rows
            ],
        }

    @staticmethod
    def create_schedule(workspace_id, report_id, frequency, delivery_channel, delivery_target):
        """Create report schedule metadata."""
        if frequency not in {'daily', 'weekly', 'monthly'}:
            raise ValueError('Invalid frequency')

        if delivery_channel != 'email':
            raise ValueError('Invalid delivery channel')

        if not delivery_target:
            raise ValueError('Delivery target is required')

        report = ReportService.get_report(report_id, workspace_id)
        if not report:
            raise ValueError('Report not found')

        schedule = ReportSchedule(
            workspace_id=workspace_id,
            report_id=report_id,
            frequency=frequency,
            delivery_channel=delivery_channel,
            delivery_target=delivery_target,
            is_active=True,
        )

        try:
            db.session.add(schedule)
            db.session.commit()
            return schedule
        except Exception as exc:
            db.session.rollback()
            logger.error('Failed to create report schedule: %s', exc)
            raise

    @staticmethod
    def list_schedules(workspace_id):
        """List report schedules for workspace."""
        rows = ReportSchedule.query.filter_by(workspace_id=workspace_id).order_by(ReportSchedule.created_at.desc()).all()
        return [
            {
                'id': row.id,
                'report_id': row.report_id,
                'frequency': row.frequency,
                'delivery_channel': row.delivery_channel,
                'delivery_target': row.delivery_target,
                'is_active': row.is_active,
            }
            for row in rows
        ]

    @staticmethod
    def export_excel(report_name, report_result):
        """Export report output as Excel bytes."""
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Report'

        sheet['A1'] = 'Report Name'
        sheet['B1'] = report_name
        sheet['A2'] = 'Generated At'
        sheet['B2'] = report_result.get('generated_at', datetime.now(timezone.utc).isoformat())

        data = report_result.get('data', {})
        sheet['A4'] = 'Key'
        sheet['B4'] = 'Value'

        row_no = 5
        if isinstance(data, dict):
            for key, value in data.items():
                if isinstance(value, (dict, list)):
                    rendered = json.dumps(value, ensure_ascii=False)
                else:
                    rendered = value
                sheet.cell(row=row_no, column=1, value=str(key))
                sheet.cell(row=row_no, column=2, value=str(rendered))
                row_no += 1

        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream.getvalue()

    @staticmethod
    def export_pdf(report_name, report_result):
        """Export report output as simple PDF bytes."""
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas

        stream = io.BytesIO()
        pdf = canvas.Canvas(stream, pagesize=A4)

        y = 800
        pdf.setFont('Helvetica-Bold', 14)
        pdf.drawString(40, y, f'Report: {report_name}')
        y -= 22

        pdf.setFont('Helvetica', 10)
        pdf.drawString(40, y, f"Generated at: {report_result.get('generated_at', datetime.now(timezone.utc).isoformat())}")
        y -= 28

        data = report_result.get('data', {})
        if isinstance(data, dict):
            for key, value in data.items():
                text = f'{key}: {json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value}'
                for chunk_start in range(0, len(text), 110):
                    pdf.drawString(40, y, text[chunk_start:chunk_start + 110])
                    y -= 14
                    if y < 60:
                        pdf.showPage()
                        pdf.setFont('Helvetica', 10)
                        y = 800

        pdf.save()
        stream.seek(0)
        return stream.getvalue()

    @staticmethod
    def serialize_report(report):
        """Serialize report row."""
        return {
            'id': report.id,
            'name': report.name,
            'report_type': report.report_type,
            'config': ReportService._safe_config(report.config_json),
            'created_by': report.created_by,
            'created_at': report.created_at.isoformat() if report.created_at else None,
            'updated_at': report.updated_at.isoformat() if report.updated_at else None,
        }

    @staticmethod
    def _safe_config(config_json):
        """Parse config json with fallback."""
        if not config_json:
            return {}
        try:
            parsed = json.loads(config_json)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}
