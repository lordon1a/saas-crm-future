from flask import Blueprint, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
import os
try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    openpyxl = None
    Font = None
    PatternFill = None
from io import BytesIO
from datetime import datetime
import csv
import json

import_bp = Blueprint('import_wizard', __name__)

# Configuration
UPLOAD_FOLDER = 'uploads/imports'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
MAX_ROWS = 100000

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _ensure_import_dependencies(require_excel=False):
    if pd is None:
        return jsonify({'error': 'Import module dependency missing: pandas'}), 503
    if require_excel and openpyxl is None:
        return jsonify({'error': 'Import module dependency missing: openpyxl'}), 503
    return None

@import_bp.route('/import')
def import_page():
    """Render the import wizard page"""
    return render_template('import.html')

@import_bp.route('/api/v1/import/template/<object_type>')
def download_template(object_type):
    """Generate and download Excel template for selected object type"""
    dep_error = _ensure_import_dependencies(require_excel=True)
    if dep_error:
        return dep_error
    
    templates = {
        'contacts': {
            'headers': ['Kişi - Ad *', 'Kişi - Soyad', 'Kişi - E-posta', 'Kişi - Telefon', 'Organizasyon - Ad', 'Kişi - Unvan', 'Kişi - Rol'],
            'example_data': [
                ['Tony', 'Turner', 'tony.turner@innovate.io', '570-608-8578', 'Moveit Limited', 'CEO', 'Decision Maker'],
                ['Hashim', 'Handy', 'hashim.handy@example.com', '740-707-3884', 'ABC Inc', 'Manager', 'Champion']
            ]
        },
        'companies': {
            'headers': ['Organizasyon - Ad *', 'Organizasyon - Adres', 'Organizasyon - Website', 'Organizasyon - Sektör'],
            'example_data': [
                ['Moveit Limited', '5,643 Finch Avenue', 'moveit.com', 'Technology'],
                ['ABC Inc', '9974 Pleasant Street', 'abc.com', 'Consulting']
            ]
        },
        'leads': {
            'headers': ['Müşteri Adayı - Başlık *', 'Kişi - Ad', 'Müşteri Adayı - Değer', 'Müşteri Adayı - Kaynak'],
            'example_data': [
                ['Yeni Proje', 'John Doe', '5000', 'Website'],
                ['Potansiyel Müşteri', 'Jane Smith', '3000', 'Referans']
            ]
        },
        'deals': {
            'headers': ['Anlaşma - Başlık *', 'Anlaşma - Değer *', 'Anlaşma - Aşama', 'Anlaşma - Kapanış Tarihi'],
            'example_data': [
                ['Büyük Satış', '10000', 'Teklif', '2024-04-15'],
                ['Yeni Müşteri', '5000', 'Görüşme', '2024-05-01']
            ]
        },
        'activities': {
            'headers': ['Etkinlik - Konu *', 'Etkinlik - Tip *', 'Etkinlik - Tarih', 'Etkinlik - Süre'],
            'example_data': [
                ['Müşteri Görüşmesi', 'Toplantı', '2024-03-20', '60'],
                ['Takip Araması', 'Arama', '2024-03-21', '15']
            ]
        },
        'products': {
            'headers': ['Ürün - Ad *', 'Ürün - Fiyat *', 'Ürün - Kod', 'Ürün - Kategori'],
            'example_data': [
                ['Premium Paket', '99', 'PRE-001', 'Yazılım'],
                ['Standart Paket', '49', 'STD-001', 'Yazılım']
            ]
        }
    }
    
    if object_type not in templates:
        return jsonify({'error': 'Invalid object type'}), 400
    
    template = templates[object_type]
    
    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = object_type.capitalize()
    
    # Header styling
    header_fill = PatternFill(start_color="3175D3", end_color="3175D3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    for col_idx, header in enumerate(template['headers'], 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
    
    # Write example data
    for row_idx, row_data in enumerate(template['example_data'], 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'{object_type}_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@import_bp.route('/api/v1/import/upload', methods=['POST'])
def upload_file():
    """Handle file upload and extract headers"""
    dep_error = _ensure_import_dependencies(require_excel=True)
    if dep_error:
        return dep_error
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    object_type = request.form.get('object_type')
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Only CSV and Excel files are allowed'}), 400
    
    # Check file size
    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)
    
    if file_size > MAX_FILE_SIZE:
        return jsonify({'error': f'File size exceeds {MAX_FILE_SIZE // (1024*1024)}MB limit'}), 400
    
    try:
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(UPLOAD_FOLDER, unique_filename)
        file.save(filepath)
        
        # Read file and extract headers
        if filename.endswith('.csv'):
            df = pd.read_csv(filepath, nrows=5)
        else:
            df = pd.read_excel(filepath, nrows=5)
        
        # Check row count
        if len(df) > MAX_ROWS:
            os.remove(filepath)
            return jsonify({'error': f'File contains more than {MAX_ROWS:,} rows'}), 400
        
        headers = df.columns.tolist()
        preview_data = df.head(3).to_dict('records')
        
        return jsonify({
            'success': True,
            'file_id': unique_filename,
            'headers': headers,
            'preview_data': preview_data,
            'row_count': len(df)
        })
        
    except Exception as e:
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500


@import_bp.route('/api/v1/import/suggest-mapping', methods=['POST'])
def suggest_mapping():
    """Advanced field mapping with multi-layer matching algorithm"""
    import re
    import unicodedata
    
    data = request.json
    file_headers = data.get('file_headers', [])
    object_type = data.get('object_type')
    preview_data = data.get('preview_data', [])
    
    # ============================================================================
    # PREPROCESSING LAYER: Text Normalization
    # ============================================================================
    def normalize_text(text):
        """
        Normalize text for consistent matching:
        1. Lowercase
        2. Remove Turkish characters (ş->s, ı->i, ğ->g, ü->u, ö->o, ç->c)
        3. Remove special characters and spaces
        """
        if not text:
            return ''
        
        # Lowercase
        text = text.lower()
        
        # Turkish character normalization
        turkish_map = {
            'ş': 's', 'ı': 'i', 'ğ': 'g', 'ü': 'u', 'ö': 'o', 'ç': 'c',
            'Ş': 's', 'İ': 'i', 'Ğ': 'g', 'Ü': 'u', 'Ö': 'o', 'Ç': 'c'
        }
        for turkish, latin in turkish_map.items():
            text = text.replace(turkish, latin)
        
        # Remove accents and diacritics
        text = unicodedata.normalize('NFKD', text)
        text = ''.join([c for c in text if not unicodedata.combining(c)])
        
        # Remove special characters, spaces, underscores, hyphens
        text = re.sub(r'[^a-z0-9]', '', text)
        
        return text
    
    # ============================================================================
    # LAYER 1: Deterministic Alias Dictionary (Sektörel Kavram Sözlüğü)
    # ============================================================================
    # Comprehensive alias mapping - normalized versions
    ALIAS_DICTIONARY = {
        'contacts': {
            'first_name': [
                'ad', 'isim', 'adi', 'firstname', 'name', 'fname', 'givenname',
                'kisiadi', 'kisiismi', 'ilkad', 'ilkisim', 'first', 'ism'
            ],
            'last_name': [
                'soyad', 'soyadi', 'lastname', 'surname', 'familyname', 'lname',
                'kisisoyadi', 'soyisim', 'last', 'soyadı'
            ],
            'email': [
                'email', 'eposta', 'mail', 'epost', 'elektronikposta', 'posta',
                'emailaddress', 'emailadresi', 'mailadresi', 'iletisim'
            ],
            'phone': [
                'telefon', 'phone', 'tel', 'mobile', 'gsm', 'cep', 'telephone',
                'telefonu', 'telno', 'telefonno', 'mobiltelefon', 'ceptelefonu',
                'iletisimno', 'numarasi', 'numara'
            ],
            'company_name': [
                'sirket', 'company', 'organizasyon', 'kurulus', 'firma', 'organization',
                'sirketadi', 'firmaadi', 'kurulusadi', 'org', 'companyname', 'kurum'
            ],
            'job_title': [
                'unvan', 'title', 'pozisyon', 'position', 'jobtitle', 'meslek',
                'isunvani', 'isunvan', 'kisiünvan', 'kisiünvani', 'kisiünvanı',
                'kisiunvan', 'kisiunvani', 'personelunvan', 'calisanunvan',
                'ispozisyonu', 'gorevunvani', 'meslekadi', 'pozisyonu'
            ],
            'role': [
                'rol', 'role', 'gorev', 'duty', 'rolunvan', 'kisirol', 'kisirolü',
                'kisirolü', 'kisirolu', 'personelrol', 'calisanrol', 'rolü',
                'rolu', 'gorevtanimi', 'sorumluluk', 'sorumluluğu', 'görevi'
            ]
        },
        'companies': {
            'name': [
                'ad', 'isim', 'name', 'sirketadi', 'companyname', 'firmaadi',
                'kurulusadi', 'sirket', 'firma', 'company'
            ],
            'address': [
                'adres', 'address', 'lokasyon', 'location', 'yer', 'addr', 'konum'
            ],
            'website': [
                'website', 'web', 'site', 'url', 'internetsitesi', 'websitesi',
                'webadres', 'link', 'websayfasi'
            ],
            'industry': [
                'sektor', 'industry', 'alan', 'sector', 'branch', 'sektorel',
                'faaliyet', 'faaliyetalani'
            ]
        }
    }
    
    if object_type not in ALIAS_DICTIONARY:
        return jsonify({'error': 'Invalid object type'}), 400
    
    # Normalize alias dictionary
    normalized_aliases = {}
    for db_field, aliases in ALIAS_DICTIONARY[object_type].items():
        normalized_aliases[db_field] = [normalize_text(alias) for alias in aliases]
    
    suggested_mapping = {}
    confidence_scores = {}
    
    # ============================================================================
    # LAYER 2: Content-Based Data Profiling (İçerik Tabanlı Profil Çıkarma)
    # ============================================================================
    def analyze_column_content(column_data, column_name):
        """
        Analyze column content to infer field type using heuristics
        Returns: (field_type, confidence, method)
        """
        if not column_data:
            return None, 0, None
        
        # Remove empty/null values
        values = [str(v).strip() for v in column_data if pd.notna(v) and str(v).strip()]
        if not values:
            return None, 0, None
        
        total = len(values)
        
        # Email detection (Regex pattern)
        email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
        email_count = sum(1 for v in values if email_pattern.match(v))
        email_ratio = email_count / total
        
        # Phone detection (10+ digits, may contain +, -, spaces, parentheses)
        phone_pattern = re.compile(r'^[\d\s\-\+\(\)]{10,}$')
        digit_heavy = [v for v in values if len(re.findall(r'\d', v)) >= 10]
        phone_count = sum(1 for v in digit_heavy if phone_pattern.match(v))
        phone_ratio = phone_count / total
        
        # URL/Website detection
        url_pattern = re.compile(r'^(https?://|www\.)[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}')
        url_count = sum(1 for v in values if url_pattern.match(v))
        url_ratio = url_count / total
        
        # Numeric detection (for prices, values)
        numeric_count = sum(1 for v in values if re.match(r'^\d+([.,]\d+)?$', v.replace(',', '')))
        numeric_ratio = numeric_count / total
        
        # Decision logic with confidence thresholds
        if email_ratio >= 0.8:
            return 'email', email_ratio, 'content_email'
        elif phone_ratio >= 0.7:
            return 'phone', phone_ratio, 'content_phone'
        elif url_ratio >= 0.7:
            return 'website', url_ratio, 'content_url'
        elif numeric_ratio >= 0.8:
            # Could be price, value, or score
            return 'numeric', numeric_ratio, 'content_numeric'
        
        return None, 0, None
    
    # ============================================================================
    # LAYER 3: Fuzzy String Matching (Levenshtein Distance)
    # ============================================================================
    def levenshtein_distance(s1, s2):
        if len(s1) < len(s2):
            return levenshtein_distance(s2, s1)
        if len(s2) == 0:
            return len(s1)
        
        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row
        
        return previous_row[-1]
    
    def similarity_ratio(s1, s2):
        distance = levenshtein_distance(s1, s2)
        max_len = max(len(s1), len(s2))
        return 1 - (distance / max_len) if max_len > 0 else 0
    
    # ============================================================================
    # MAIN MATCHING FLOW
    # ============================================================================
    for idx, file_header in enumerate(file_headers):
        # Normalize file header
        normalized_header = normalize_text(file_header)
        
        best_match = None
        best_score = 0
        match_method = None
        
        # STEP 1: Exact match in alias dictionary (Deterministic)
        for db_field, aliases in normalized_aliases.items():
            if normalized_header in aliases:
                best_match = db_field
                best_score = 1.0
                match_method = 'exact_alias'
                break
        
        # STEP 2: Content-based profiling (if no exact match)
        if best_score < 1.0 and preview_data:
            column_data = [row.get(file_header) for row in preview_data if file_header in row]
            inferred_type, content_confidence, content_method = analyze_column_content(column_data, file_header)
            
            if inferred_type and content_confidence >= 0.7:
                # Map inferred type to db field
                if inferred_type in normalized_aliases:
                    best_match = inferred_type
                    best_score = content_confidence
                    match_method = content_method
        
        # STEP 3: Fuzzy string matching (last resort)
        if best_score < 0.85:
            for db_field, aliases in normalized_aliases.items():
                for alias in aliases:
                    ratio = similarity_ratio(normalized_header, alias)
                    
                    # High confidence threshold (85%)
                    if ratio >= 0.85 and ratio > best_score:
                        best_score = ratio
                        best_match = db_field
                        match_method = 'fuzzy_match'
        
        # Store mapping if confidence is sufficient (>60%)
        if best_match and best_score > 0.6:
            suggested_mapping[file_header] = best_match
            confidence_scores[file_header] = {
                'score': round(best_score * 100, 1),
                'method': match_method,
                'normalized': normalized_header
            }
    
    # Calculate overall confidence
    overall_confidence = len(suggested_mapping) / len(file_headers) if file_headers else 0
    
    return jsonify({
        'suggested_mapping': suggested_mapping,
        'confidence_scores': confidence_scores,
        'overall_confidence': round(overall_confidence * 100, 1),
        'total_fields': len(file_headers),
        'mapped_fields': len(suggested_mapping)
    })


@import_bp.route('/api/v1/import/validate', methods=['POST'])
def validate_import():
    """Validate import data before processing"""
    dep_error = _ensure_import_dependencies(require_excel=True)
    if dep_error:
        return dep_error
    from models_crm import Contact, Company
    from models import db
    from flask import session
    
    data = request.json
    file_id = data.get('file_id')
    field_mapping = data.get('field_mapping')
    object_type = data.get('object_type')
    
    workspace_id = session.get('workspace_id', 1)
    filepath = os.path.join(UPLOAD_FOLDER, file_id)
    
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    
    try:
        # Read file
        if file_id.endswith('.csv'):
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)
        
        errors = []
        warnings = []
        duplicates = []
        
        # Validate required fields
        required_fields = {
            'contacts': ['first_name'],
            'companies': ['name'],
            'leads': ['title'],
            'deals': ['title', 'value'],
            'activities': ['subject', 'type'],
            'products': ['name', 'price']
        }
        
        for required_field in required_fields.get(object_type, []):
            if required_field not in field_mapping.values():
                errors.append(f'Required field "{required_field}" is not mapped')
        
        # Check for duplicates in database
        if object_type == 'contacts':
            for idx, row in df.iterrows():
                row_num = idx + 2
                
                # Check for empty required fields
                for file_col, db_col in field_mapping.items():
                    if db_col in required_fields.get(object_type, []):
                        if pd.isna(row[file_col]) or str(row[file_col]).strip() == '':
                            errors.append(f'Row {row_num}: Required field "{db_col}" is empty')
                
                # Get contact data
                contact_data = {}
                for file_col, db_col in field_mapping.items():
                    value = row[file_col]
                    if pd.notna(value):
                        contact_data[db_col] = str(value).strip()
                
                # Check for duplicates by email or name
                email = contact_data.get('email')
                first_name = contact_data.get('first_name', '')
                last_name = contact_data.get('last_name', '')
                
                existing_contacts = []
                
                if email:
                    existing = Contact.query.filter_by(
                        workspace_id=workspace_id,
                        email=email,
                        is_deleted=False
                    ).first()
                    if existing:
                        existing_contacts.append({
                            'id': existing.id,
                            'match_type': 'email',
                            'existing': {
                                'name': f"{existing.first_name} {existing.last_name}",
                                'email': existing.email,
                                'phone': existing.phone,
                                'company': existing.company.name if existing.company else None
                            }
                        })
                
                # Check by name if no email match
                if not existing_contacts and first_name:
                    existing = Contact.query.filter_by(
                        workspace_id=workspace_id,
                        first_name=first_name,
                        last_name=last_name,
                        is_deleted=False
                    ).first()
                    if existing:
                        existing_contacts.append({
                            'id': existing.id,
                            'match_type': 'name',
                            'existing': {
                                'name': f"{existing.first_name} {existing.last_name}",
                                'email': existing.email,
                                'phone': existing.phone,
                                'company': existing.company.name if existing.company else None
                            }
                        })
                
                if existing_contacts:
                    duplicates.append({
                        'row': row_num,
                        'new_data': {
                            'name': f"{first_name} {last_name}",
                            'email': email,
                            'phone': contact_data.get('phone'),
                            'company': contact_data.get('company_name')
                        },
                        'matches': existing_contacts
                    })
                
                # Validate email format
                if email and '@' not in email:
                    warnings.append(f'Row {row_num}: Invalid email format "{email}"')
                
                # Validate phone format
                phone = contact_data.get('phone')
                if phone and len(phone.replace('-', '').replace(' ', '').replace('+', '')) < 10:
                    warnings.append(f'Row {row_num}: Phone number may be invalid "{phone}"')
        
        elif object_type == 'companies':
            for idx, row in df.iterrows():
                row_num = idx + 2
                
                # Check for empty required fields
                for file_col, db_col in field_mapping.items():
                    if db_col in required_fields.get(object_type, []):
                        if pd.isna(row[file_col]) or str(row[file_col]).strip() == '':
                            errors.append(f'Row {row_num}: Required field "{db_col}" is empty')
                
                # Get company name
                name_col = [k for k, v in field_mapping.items() if v == 'name'][0]
                company_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
                
                if company_name:
                    existing = Company.query.filter_by(
                        workspace_id=workspace_id,
                        name=company_name,
                        is_deleted=False
                    ).first()
                    
                    if existing:
                        duplicates.append({
                            'row': row_num,
                            'new_data': {'name': company_name},
                            'matches': [{
                                'id': existing.id,
                                'match_type': 'name',
                                'existing': {
                                    'name': existing.name,
                                    'address': existing.address,
                                    'website': existing.website
                                }
                            }]
                        })
        
        else:
            # Validate other object types
            for idx, row in df.iterrows():
                row_num = idx + 2
                
                for file_col, db_col in field_mapping.items():
                    if db_col in required_fields.get(object_type, []):
                        if pd.isna(row[file_col]) or str(row[file_col]).strip() == '':
                            errors.append(f'Row {row_num}: Required field "{db_col}" is empty')
        
        return jsonify({
            'valid': len(errors) == 0,
            'errors': errors[:50],
            'warnings': warnings[:50],
            'duplicates': duplicates[:50],
            'total_rows': len(df),
            'duplicate_count': len(duplicates)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Validation error: {str(e)}'}), 500


@import_bp.route('/api/v1/import/execute', methods=['POST'])
def execute_import():
    """Execute the import process"""
    dep_error = _ensure_import_dependencies(require_excel=True)
    if dep_error:
        return dep_error
    from models_crm import Contact, Company, CustomField, CustomFieldValue
    from models import db
    from flask import session
    
    data = request.json
    file_id = data.get('file_id')
    field_mapping = data.get('field_mapping')
    object_type = data.get('object_type')
    duplicate_action = data.get('duplicate_action', 'skip')  # skip, update, create, or create_with_suffix
    
    # Get workspace_id from session
    workspace_id = session.get('workspace_id', 1)  # Default to 1 if not in session
    
    filepath = os.path.join(UPLOAD_FOLDER, file_id)
    
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    
    try:
        # Read file
        if file_id.endswith('.csv'):
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)
        
        imported_count = 0
        updated_count = 0
        skipped_count = 0
        failed_count = 0
        errors = []
        
        # Define standard fields for each object type
        STANDARD_FIELDS = {
            'contacts': ['first_name', 'last_name', 'email', 'phone', 'company_name', 'job_title', 'role', 'lead_score'],
            'companies': ['name', 'address', 'website', 'industry']
        }
        
        # Identify unmapped columns (potential custom fields)
        all_file_columns = set(df.columns)
        mapped_columns = set(field_mapping.keys())
        unmapped_columns = all_file_columns - mapped_columns
        
        # Create custom fields for unmapped columns
        custom_field_map = {}
        if unmapped_columns:
            for col_name in unmapped_columns:
                # Check if custom field already exists
                custom_field = CustomField.query.filter_by(
                    workspace_id=workspace_id,
                    entity_type=object_type.rstrip('s'),  # 'contacts' -> 'contact'
                    field_name=col_name
                ).first()
                
                if not custom_field:
                    # Create new custom field
                    custom_field = CustomField(
                        workspace_id=workspace_id,
                        entity_type=object_type.rstrip('s'),
                        field_name=col_name,
                        field_type='text',  # Default to text, can be enhanced later
                        is_required=False
                    )
                    db.session.add(custom_field)
                    db.session.flush()
                
                custom_field_map[col_name] = custom_field.id
        
        # Import based on object type
        if object_type == 'contacts':
            for idx, row in df.iterrows():
                try:
                    # Map fields
                    contact_data = {}
                    for file_col, db_col in field_mapping.items():
                        value = row[file_col]
                        if pd.notna(value):
                            contact_data[db_col] = str(value).strip()
                    
                    # Handle name field (split if needed)
                    if 'first_name' in contact_data:
                        first_name = contact_data['first_name']
                        last_name = contact_data.get('last_name', '')
                    elif 'name' in contact_data:
                        # Split full name
                        name_parts = contact_data['name'].split(' ', 1)
                        first_name = name_parts[0]
                        last_name = name_parts[1] if len(name_parts) > 1 else ''
                    else:
                        raise ValueError('Name field is required')
                    
                    # Handle company
                    company_id = None
                    if 'company_name' in contact_data:
                        company_name = contact_data['company_name']
                        # Find or create company
                        company = Company.query.filter_by(
                            name=company_name,
                            workspace_id=workspace_id,
                            is_deleted=False
                        ).first()
                        if not company:
                            company = Company(
                                name=company_name,
                                workspace_id=workspace_id,
                                is_deleted=False
                            )
                            db.session.add(company)
                            db.session.flush()
                        company_id = company.id
                    
                    # Check for duplicate contact
                    # Email varsa sadece email, yoksa sadece isim ile kontrol
                    email = contact_data.get('email')
                    existing_contact = None
                    
                    if email:
                        # Email varsa SADECE email ile duplicate kontrolü yap
                        existing_contact = Contact.query.filter_by(
                            workspace_id=workspace_id,
                            email=email,
                            is_deleted=False
                        ).first()
                    elif first_name:
                        # Email YOKSA SADECE isim ile duplicate kontrolü yap
                        existing_contact = Contact.query.filter_by(
                            workspace_id=workspace_id,
                            first_name=first_name,
                            last_name=last_name,
                            is_deleted=False
                        ).first()
                    
                    # Handle duplicate based on action
                    if existing_contact:
                        if duplicate_action == 'skip':
                            skipped_count += 1
                            continue
                        elif duplicate_action == 'update':
                            # Update existing contact
                            existing_contact.first_name = first_name
                            existing_contact.last_name = last_name
                            if email:
                                existing_contact.email = email
                            if contact_data.get('phone'):
                                existing_contact.phone = contact_data.get('phone')
                            if company_id:
                                existing_contact.company_id = company_id
                            if contact_data.get('job_title'):
                                existing_contact.job_title = contact_data.get('job_title')
                            if contact_data.get('role'):
                                existing_contact.role = contact_data.get('role')
                            
                            # Update custom fields
                            for col_name, custom_field_id in custom_field_map.items():
                                if col_name in row.index and pd.notna(row[col_name]):
                                    custom_value = CustomFieldValue.query.filter_by(
                                        custom_field_id=custom_field_id,
                                        entity_id=existing_contact.id
                                    ).first()
                                    if custom_value:
                                        custom_value.value = str(row[col_name]).strip()
                                    else:
                                        custom_value = CustomFieldValue(
                                            custom_field_id=custom_field_id,
                                            entity_id=existing_contact.id,
                                            value=str(row[col_name]).strip()
                                        )
                                        db.session.add(custom_value)
                            
                            updated_count += 1
                            continue
                        elif duplicate_action == 'create_with_suffix':
                            # Create new contact with modified email/name to avoid duplicates
                            suffix = 1
                            original_email = email
                            original_first_name = first_name
                            
                            # Find unique email
                            if email:
                                email_parts = email.rsplit('@', 1)
                                while Contact.query.filter_by(workspace_id=workspace_id, email=email).first():
                                    if len(email_parts) == 2:
                                        email = f"{email_parts[0]}{suffix}@{email_parts[1]}"
                                    else:
                                        email = f"{original_email}{suffix}"
                                    suffix += 1
                            
                            # Find unique name if no email
                            if not original_email:
                                suffix = 1
                                while Contact.query.filter_by(
                                    workspace_id=workspace_id,
                                    first_name=first_name,
                                    last_name=last_name
                                ).first():
                                    first_name = f"{original_first_name} ({suffix})"
                                    suffix += 1
                            
                            # Continue to create with modified data
                        # else: duplicate_action == 'create' - continue to create new (may cause duplicates)
                    
                    # Create contact
                    contact = Contact(
                        workspace_id=workspace_id,
                        first_name=first_name,
                        last_name=last_name,
                        email=contact_data.get('email') if not existing_contact or duplicate_action != 'create_with_suffix' else email,
                        phone=contact_data.get('phone'),
                        company_id=company_id,
                        job_title=contact_data.get('job_title'),
                        role=contact_data.get('role'),
                        lead_score=int(contact_data.get('lead_score', 0)) if contact_data.get('lead_score') else 0
                    )
                    
                    db.session.add(contact)
                    db.session.flush()  # Get contact ID
                    
                    # Add custom field values
                    for col_name, custom_field_id in custom_field_map.items():
                        if col_name in row.index and pd.notna(row[col_name]):
                            custom_value = CustomFieldValue(
                                custom_field_id=custom_field_id,
                                entity_id=contact.id,
                                value=str(row[col_name]).strip()
                            )
                            db.session.add(custom_value)
                    
                    imported_count += 1
                    
                except Exception as e:
                    failed_count += 1
                    errors.append(f'Row {idx + 2}: {str(e)}')
                    continue
            
            db.session.commit()
        
        elif object_type == 'companies':
            for idx, row in df.iterrows():
                try:
                    contact_data = {}
                    for file_col, db_col in field_mapping.items():
                        value = row[file_col]
                        if pd.notna(value):
                            contact_data[db_col] = str(value).strip()
                    
                    if 'name' not in contact_data:
                        raise ValueError('Company name is required')
                    
                    # Check for duplicate
                    existing = Company.query.filter_by(
                        name=contact_data['name'],
                        workspace_id=workspace_id,
                        is_deleted=False
                    ).first()
                    
                    if existing:
                        if duplicate_action == 'skip':
                            skipped_count += 1
                            continue
                        elif duplicate_action == 'update':
                            # Update existing company
                            if contact_data.get('address'):
                                existing.address = contact_data.get('address')
                            if contact_data.get('website'):
                                existing.website = contact_data.get('website')
                            if contact_data.get('industry'):
                                existing.industry = contact_data.get('industry')
                            
                            # Update custom fields
                            for col_name, custom_field_id in custom_field_map.items():
                                if col_name in row.index and pd.notna(row[col_name]):
                                    custom_value = CustomFieldValue.query.filter_by(
                                        custom_field_id=custom_field_id,
                                        entity_id=existing.id
                                    ).first()
                                    if custom_value:
                                        custom_value.value = str(row[col_name]).strip()
                                    else:
                                        custom_value = CustomFieldValue(
                                            custom_field_id=custom_field_id,
                                            entity_id=existing.id,
                                            value=str(row[col_name]).strip()
                                        )
                                        db.session.add(custom_value)
                            
                            updated_count += 1
                            continue
                        elif duplicate_action == 'create_with_suffix':
                            # Create new company with modified name to avoid duplicates
                            suffix = 1
                            original_name = contact_data['name']
                            company_name = original_name
                            
                            while Company.query.filter_by(workspace_id=workspace_id, name=company_name, is_deleted=False).first():
                                company_name = f"{original_name} ({suffix})"
                                suffix += 1
                            
                            contact_data['name'] = company_name
                            # Continue to create with modified name
                        # else: duplicate_action == 'create' - continue to create new
                    
                    company = Company(
                        workspace_id=workspace_id,
                        name=contact_data['name'],
                        address=contact_data.get('address'),
                        website=contact_data.get('website'),
                        industry=contact_data.get('industry')
                    )
                    
                    db.session.add(company)
                    db.session.flush()  # Get company ID
                    
                    # Add custom field values
                    for col_name, custom_field_id in custom_field_map.items():
                        if col_name in row.index and pd.notna(row[col_name]):
                            custom_value = CustomFieldValue(
                                custom_field_id=custom_field_id,
                                entity_id=company.id,
                                value=str(row[col_name]).strip()
                            )
                            db.session.add(custom_value)
                    
                    imported_count += 1
                    
                except Exception as e:
                    failed_count += 1
                    errors.append(f'Row {idx + 2}: {str(e)}')
                    continue
            
            db.session.commit()
        
        # Clean up uploaded file
        try:
            os.remove(filepath)
        except:
            pass
        
        import_job_id = f"import_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        return jsonify({
            'success': True,
            'job_id': import_job_id,
            'message': 'Import completed',
            'imported_rows': imported_count,
            'updated_rows': updated_count,
            'skipped_rows': skipped_count,
            'failed_rows': failed_count,
            'total_rows': len(df),
            'errors': errors[:10]  # Return first 10 errors
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Import error: {str(e)}'}), 500


@import_bp.route('/api/v1/import/status/<job_id>')
def import_status(job_id):
    """Check import job status"""
    
    # Since we're doing synchronous import, return completed status
    # In production with Celery, this would check actual task status
    
    return jsonify({
        'job_id': job_id,
        'status': 'completed',
        'progress': 100
    })
